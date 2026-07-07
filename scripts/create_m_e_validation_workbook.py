import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Directories
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
DOCS_DIR = os.path.join(BASE_DIR, 'docs', 'registry-validation')

# Mapping of sheet names to CSV filenames
sheets = {
    'IP Registry': 'ip_registry_validation.csv',
    'Activity Registry': 'activity_registry_validation.csv',
    'Indicator Registry': 'indicator_registry_validation.csv',
    'Target Registry': 'target_registry_validation.csv',
    'Activity Indicator Crosswalk': 'activity_indicator_crosswalk_validation.csv',
    'Evidence Requirement': 'evidence_requirement_validation.csv',
    'NRCS PeaceWin Gap Review': 'nrcs_peacewin_gap_review.csv',
    'Unresolved Mapping Review': 'unresolved_mapping_review.csv',
    'BigQuery Join Readiness': 'bigquery_join_readiness_review.csv',
    'QA Summary - Instructions': 'REGISTRY_VALIDATION_GUIDE.md'  # using guide as a text tab
}

workbook_path = os.path.join(DOCS_DIR, 'UNFPA_MEL_Registry_M_E_Validation_Package.xlsx')

# Create Excel writer
with pd.ExcelWriter(workbook_path, engine='openpyxl') as writer:
    for sheet_name, filename in sheets.items():
        file_path = os.path.join(DOCS_DIR, filename)
        if filename.lower().endswith('.csv'):
            df = pd.read_csv(file_path, dtype=str).fillna('')
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # For the guide markdown, insert as a single cell
            ws = writer.book.create_sheet(title=sheet_name)
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            ws.cell(row=1, column=1, value=content)
    # Save the workbook before formatting


# Apply formatting
wb = load_workbook(workbook_path)
me_review_count = 0
nrcs_gap_count = 0
unknown_evidence_count = 0

yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Freeze header row
    ws.freeze_panes = ws['A2']
    # Apply filter to all columns if there is data
    max_col = ws.max_column
    max_row = ws.max_row
    if max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    # Auto-size columns
    for col in range(1, max_col + 1):
        column = get_column_letter(col)
        max_length = 0
        for cell in ws[column]:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    # Highlight rows based on specific columns
    header = [cell.value for cell in ws[1]]
    if 'me_review_status' in header:
        idx = header.index('me_review_status') + 1
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=idx)
            if cell.value and cell.value.strip():
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill
                me_review_count += 1
    if 'mapping_status' in header:
        idx = header.index('mapping_status') + 1
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=idx)
            if cell.value and cell.value.strip() == 'incomplete_source':
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = red_fill
                nrcs_gap_count += 1
    if 'evidence_status' in header:
        idx = header.index('evidence_status') + 1
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=idx)
            if cell.value and cell.value.lower() == 'unknown':
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = red_fill
                unknown_evidence_count += 1
    if 'evidence_status' in header:
        idx = header.index('evidence_status') + 1
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=idx)
            if cell.value and cell.value.lower() == 'unknown':
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = red_fill
                unknown_evidence_count += 1

wb.save(workbook_path)

summary = {
    "workbook_path": workbook_path,
    "sheets": len(sheets),
    "me_review_records": me_review_count,
    "nrcs_gap_records": nrcs_gap_count,
    "unknown_evidence_records": unknown_evidence_count
}
print(json.dumps(summary))
